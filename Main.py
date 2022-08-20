from openpyxl import Workbook
from Excel import Excel_1
from DataBase import Database_1
import xlwt
import openpyxl
import pyexcel
from openpyxl.styles import PatternFill
import time
from Ruta import *
import sys
import msvcrt

#Clase principal   
if __name__ == "__main__":
    
    numGrupo1 = [1,2,3,4,5,6,7,8,9]
    numGrupo2 = [10,11,12,13,14,15,16,17,18]
    numGrupo3 = [19,20,21,22,23,24,25,26,27]
    numGrupo4 = [28,29,30,31,32,33,34,35,36]
    numGrupo5 = [37,38,39,40,41,42,43,44,45,46]
    numGrupo6 = [47,48,49]
    numGrupo7 = [51,52,53]
    conexion = Database_1()
    consulta = conexion.consulta()
    ruta = ruta()
    dia = time.strftime("%d")
    mes = time.strftime("%m")
    ano = time.strftime("%Y")
    ruta.crear_carpetas(dia=dia, mes=mes, ano=ano)

 


    
    excel = Excel_1("Grupo 1")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo1:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 1.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 1.xlsx')
    
    
    excel = Excel_1("Grupo 2")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo2:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 2.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 2.xlsx')
    
    excel = Excel_1("Grupo 3")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo3:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 3.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 3.xlsx')
    
    excel = Excel_1("Grupo 4")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo4:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 4.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 4.xlsx')
            
    excel = Excel_1("Grupo 5")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo5:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 5.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 5.xlsx')
    
    excel = Excel_1("Grupo 6")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo6:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)

    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 6.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 6.xlsx')            
    
    excel = Excel_1("Grupo 7")
    wout = xlwt.Workbook()
    excel.crear_xls(wb=wout,dia=dia,mes=ruta.dife_fecha(),ano=ano)
    for ct in range(len(consulta)):
        numero = int(str(consulta[ct][9])[9:])
        print(consulta[ct])
        if numero in numGrupo7:
            excel.escribir_xls(dato=consulta[ct],dia=dia,mes=ruta.dife_fecha(),ano=ano)
    excel.convertir(dia=dia,mes=ruta.dife_fecha(),ano=ano)
    informe = openpyxl.load_workbook(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 7.xlsx')
    sheet = informe.active 
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 24
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 24
    sheet.column_dimensions['G'].width = 14
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['D'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    sheet.column_dimensions['F'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
    informe.save(f'.\\{ano}\\{ruta.dife_fecha()}\\{dia}\\Grupo 7.xlsx')
            
    print('\n----------------------\n')  
    print("Finalizó con exito!!\nPulsa una tecla para cerrar")
    print('\n----------------------\n') 
    msvcrt.getch()
    sys.exit(1)            
